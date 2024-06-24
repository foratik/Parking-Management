import subprocess
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

iverilog_cmd = ["iverilog", "-o", "out", "Code/ParkingManagement.v"]
subprocess.run(iverilog_cmd, check=True)

vvp_cmd = ["vvp", "out"]
result = subprocess.run(vvp_cmd, capture_output=True, text=True, check=True)

output_lines = result.stdout.splitlines()
header = output_lines[0].split('\t')
data = [line.split('\t') for line in output_lines[1:]]

df = pd.DataFrame(data, columns=header)
excel_filename = "parking_management.xlsx"

with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    sheet = workbook.active

    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    even_row_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    odd_row_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    for cell in sheet[1]:
        cell.fill = header_fill

    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column), start=1):
        fill = even_row_fill if i % 2 == 0 else odd_row_fill
        for cell in row:
            cell.fill = fill

    for col in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 17

print("Output has been written to parking_management.xlsx")

import os
os.startfile(excel_filename)
